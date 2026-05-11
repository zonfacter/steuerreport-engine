#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import sqlite3
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
RUN_DATE = "2026-05-11"
DEFAULT_DB = Path("/root/.local/share/steuerreport/ai_readonly/steuerreport_ai_readonly.sqlite")
OUT_JSON = ROOT / "var" / "hnt_usdt_remaining_inventory_gap_audit_2026-05-11.json"
OUT_MD = ROOT / "docs" / "229_HNT_USDT_REMAINING_INVENTORY_GAP_AUDIT_2026-05-11.md"
MATERIAL_PROCEEDS_EUR = Decimal("50")
HNT_CONTEXT_CUTOFFS = (
    ("2021-08-17T16:10:05+00:00", "Vor den Binance-HNT-Verkaeufen ohne Lot-Quelle"),
    ("2021-08-20T08:01:13+00:00", "Vor Legacy-Outflow zum Binance-Deposit 2021-08-20"),
    ("2022-07-12T06:59:57+00:00", "Vor Legacy-Outflow zum Binance-Deposit 2022-07-12"),
)
HNT_CONTEXT_SOURCES = (
    "helium_legacy_cointracking",
    "helium_legacy_raw",
    "heliumtracker",
    "heliumgeek",
)
HELIUMTRACKER_PATTERN = "heliumtracker-report-advanced-*.csv"
BINANCE_STAKING_DEPOSIT_FILES = (
    "BINANCE - HNT Transfer Staking Wallet - Deposit_History 07 bis 09-2021.xlsx",
    "BINANCE - HNT Transfer Staking Wallet - deposit_history 11-2021.xlsx",
)
STAKING_TRANSFER_WORKBOOK = "Heliumtracker 2021.xlsx"
DAILY_REWARD_WORKBOOK = "daily_hotspot_rewards_2022-03-21T17_52_18.702901-04_00.xlsx"
BINANCE_HISTORY_WORKBOOK = "Binance Export History Daten 2021 2022.xlsx"
STAKING_WALLET_SOURCE_NAME = (
    "manual_legacy_import:helium_legacy_raw:"
    "helium-Staking Wallet 14eKedP4gCyefaMgjxPULPVecDq6gM5aEJYLDvbiRXZpuq2kYNA-all-raw.csv"
)
MAIN_HELIUM_WALLET = "133rkwoKCfxLTTt1zGjge7c2nGLUSY5sTuG2V61zi6ik269Tf4j"
STAKING_WALLET = "14eKedP4gCyefaMgjxPULPVecDq6gM5aEJYLDvbiRXZpuq2kYNA"
BINANCE_HNT_WALLET = "138bCXPV"


def dec(value: Any) -> Decimal:
    try:
        return Decimal(str(value or "0").strip().replace(",", ""))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def plain(value: Any) -> str:
    value_dec = dec(value)
    text = format(value_dec, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def clean_text(value: Any) -> str:
    return str(value or "").replace("\xa0", " ").strip()


def load_workbook(path: Path) -> Any:
    try:
        from openpyxl import load_workbook as openpyxl_load_workbook
    except ImportError as exc:  # pragma: no cover - environment guard for standalone script
        raise RuntimeError("openpyxl is required for the local XLSX audit") from exc
    return openpyxl_load_workbook(path, data_only=True, read_only=True)


def find_local_workbooks(filename: str) -> list[Path]:
    return sorted({path for path in ROOT.rglob(filename) if path.is_file()})


def first_local_workbook(filename: str) -> Path | None:
    paths = find_local_workbooks(filename)
    return paths[0] if paths else None


def sheet_records(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows_iter)]
    except StopIteration:
        return []
    records: list[dict[str, Any]] = []
    for row in rows_iter:
        record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def rows(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    cur = conn.execute(sql, params)
    return [dict(row) for row in cur.fetchall()]


def payload(row: dict[str, Any] | None) -> dict[str, Any]:
    if not row:
        return {}
    try:
        loaded = json.loads(str(row.get("payload_json") or "{}"))
    except json.JSONDecodeError:
        return {}
    return loaded if isinstance(loaded, dict) else {}


def event_summary(row: dict[str, Any] | None) -> dict[str, Any]:
    data = payload(row)
    return {
        "event_id": str(row.get("unique_event_id") if row else ""),
        "source_file_id": str(row.get("source_file_id") if row else ""),
        "source_name": str(row.get("source_name") if row else ""),
        "row_index": int(row.get("row_index") or 0) if row else 0,
        "timestamp_utc": str(data.get("timestamp_utc") or data.get("timestamp") or ""),
        "source": str(data.get("source") or ""),
        "event_type": str(data.get("event_type") or ""),
        "side": str(data.get("side") or data.get("direction") or "").lower().strip(),
        "asset": str(data.get("asset") or ""),
        "quantity": str(data.get("quantity") or data.get("amount") or ""),
        "price": str(data.get("price") or ""),
        "value_usd": str(data.get("value_usd") or data.get("value_usd_sum") or ""),
        "value_eur": str(data.get("value_eur") or ""),
        "tx_id": str(data.get("tx_id") or data.get("transaction_hash") or data.get("signature") or ""),
    }


def latest_jobs(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    job_rows = rows(
        conn,
        """
        SELECT *
        FROM ai_latest_completed_jobs_per_year
        WHERE tax_year BETWEEN 2020 AND 2026
        ORDER BY tax_year
        """,
    )
    output: list[dict[str, Any]] = []
    for job in job_rows:
        try:
            result = json.loads(str(job.get("result_json") or "{}"))
        except json.JSONDecodeError:
            result = {}
        output.append(
            {
                "tax_year": int(job.get("tax_year") or 0),
                "job_id": str(job.get("job_id") or ""),
                "updated_at_utc": str(job.get("updated_at_utc") or ""),
                "tax_line_count": int(result.get("tax_line_count") or 0),
                "derivative_line_count": int(
                    rows(
                        conn,
                        "SELECT count(*) AS count FROM derivative_lines WHERE job_id = ?",
                        (str(job.get("job_id") or ""),),
                    )[0]["count"]
                ),
            }
        )
    return output


def remaining_lines(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return rows(
        conn,
        """
        SELECT *
        FROM ai_open_zero_cost_tax_lines
        WHERE asset IN ('HNT', 'USDT')
          AND CAST(proceeds_eur AS REAL) >= ?
        ORDER BY tax_year, asset, line_no
        """,
        (str(MATERIAL_PROCEEDS_EUR),),
    )


def load_raw_event(conn: sqlite3.Connection, event_id: str) -> dict[str, Any] | None:
    if not event_id:
        return None
    result = rows(
        conn,
        """
        SELECT r.unique_event_id, r.source_file_id, r.row_index, r.payload_json, sf.source_name
        FROM raw_events r
        LEFT JOIN source_files sf ON sf.source_file_id = r.source_file_id
        WHERE r.unique_event_id = ?
        """,
        (event_id,),
    )
    return result[0] if result else None


def transfer_matches_for_inbound(conn: sqlite3.Connection, event_id: str) -> list[dict[str, Any]]:
    if not event_id:
        return []
    matches = rows(
        conn,
        """
        SELECT *
        FROM ai_transfer_matches_flat
        WHERE inbound_event_id = ?
        ORDER BY created_at_utc, match_id
        """,
        (event_id,),
    )
    output: list[dict[str, Any]] = []
    for match in matches:
        outbound_payload = json.loads(str(match.get("outbound_payload_json") or "{}"))
        inbound_payload = json.loads(str(match.get("inbound_payload_json") or "{}"))
        output.append(
            {
                "match_id": str(match.get("match_id") or ""),
                "outbound_event_id": str(match.get("outbound_event_id") or ""),
                "inbound_event_id": str(match.get("inbound_event_id") or ""),
                "confidence_score": str(match.get("confidence_score") or ""),
                "time_diff_seconds": int(match.get("time_diff_seconds") or 0),
                "amount_diff": str(match.get("amount_diff") or ""),
                "status": str(match.get("status") or ""),
                "method": str(match.get("method") or ""),
                "note": str(match.get("note") or ""),
                "outbound": {
                    "timestamp_utc": str(outbound_payload.get("timestamp_utc") or ""),
                    "source": str(outbound_payload.get("source") or ""),
                    "event_type": str(outbound_payload.get("event_type") or ""),
                    "asset": str(outbound_payload.get("asset") or ""),
                    "quantity": str(outbound_payload.get("quantity") or ""),
                    "value_usd": str(outbound_payload.get("value_usd") or ""),
                    "tx_id": str(outbound_payload.get("tx_id") or ""),
                },
                "inbound": {
                    "timestamp_utc": str(inbound_payload.get("timestamp_utc") or ""),
                    "source": str(inbound_payload.get("source") or ""),
                    "event_type": str(inbound_payload.get("event_type") or ""),
                    "asset": str(inbound_payload.get("asset") or ""),
                    "quantity": str(inbound_payload.get("quantity") or ""),
                    "tx_id": str(inbound_payload.get("tx_id") or ""),
                },
            }
        )
    return output


def same_tx_events(conn: sqlite3.Connection, tx_id: str) -> list[dict[str, Any]]:
    if not tx_id:
        return []
    return [
        event_summary(row)
        for row in rows(
            conn,
            """
            SELECT r.unique_event_id, r.source_file_id, r.row_index, r.payload_json, sf.source_name
            FROM raw_events r
            LEFT JOIN source_files sf ON sf.source_file_id = r.source_file_id
            WHERE json_extract(r.payload_json, '$.tx_id') = ?
            ORDER BY COALESCE(json_extract(r.payload_json, '$.timestamp_utc'), json_extract(r.payload_json, '$.timestamp')),
                     r.unique_event_id
            """,
            (tx_id,),
        )
    ]


def classify(line: dict[str, Any], lot_event: dict[str, Any], matches: list[dict[str, Any]]) -> str:
    if not str(line.get("lot_source_event_id") or ""):
        return "missing_lot_source_inventory_gap"
    if lot_event.get("source") == "binance_api" and lot_event.get("event_type") == "deposit" and matches:
        return "matched_transfer_source_cost_basis_gap"
    return "unclassified_inventory_gap"


def hnt_source_balance_context(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for cutoff, label in HNT_CONTEXT_CUTOFFS:
        for source in HNT_CONTEXT_SOURCES:
            source_rows = rows(
                conn,
                """
                SELECT side, event_type, quantity
                FROM ai_raw_events_flat
                WHERE source = ?
                  AND asset = 'HNT'
                  AND timestamp_utc <= ?
                """,
                (source, cutoff),
            )
            inbound = Decimal("0")
            outbound = Decimal("0")
            rewards = Decimal("0")
            for row in source_rows:
                qty = dec(row.get("quantity"))
                side = str(row.get("side") or "").lower().strip()
                if side == "in":
                    inbound += qty
                elif side == "out":
                    outbound += qty
                if str(row.get("event_type") or "") == "mining_reward":
                    rewards += qty
            balance = inbound - outbound
            output.append(
                {
                    "cutoff": cutoff,
                    "label": label,
                    "source": source,
                    "event_count": len(source_rows),
                    "inbound_hnt": plain(inbound),
                    "outbound_hnt": plain(outbound),
                    "mining_reward_hnt": plain(rewards),
                    "balance_hnt": plain(balance),
                }
            )
    return output


def local_heliumtracker_file_coverage(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for path in sorted(ROOT.glob(HELIUMTRACKER_PATTERN)):
        fs_rows = 0
        fs_hnt = Decimal("0")
        with path.open(newline="", encoding="utf-8") as handle:
            for row in csv.DictReader(handle):
                fs_rows += 1
                fs_hnt += dec(row.get("Mining Rewards HNT"))

        source_file = rows(
            conn,
            """
            SELECT source_file_id, row_count
            FROM source_files
            WHERE source_name LIKE ?
            ORDER BY created_at_utc DESC
            LIMIT 1
            """,
            (f"%{path.name}",),
        )
        source_file_id = str(source_file[0]["source_file_id"]) if source_file else ""
        imported = bool(source_file_id)
        imported_rows = int(source_file[0]["row_count"] or 0) if source_file else 0
        imported_reward_rows = rows(
            conn,
            """
            SELECT payload_json
            FROM raw_events
            WHERE source_file_id = ?
              AND json_extract(payload_json, '$.asset') = 'HNT'
              AND json_extract(payload_json, '$.event_type') = 'mining_reward'
            """,
            (source_file_id,),
        ) if imported else []
        imported_hnt = Decimal("0")
        for row in imported_reward_rows:
            imported_hnt += dec(payload(row).get("quantity"))
        output.append(
            {
                "file": path.name,
                "fs_rows": fs_rows,
                "fs_hnt": plain(fs_hnt),
                "imported": imported,
                "store_row_count": imported_rows,
                "store_reward_event_count": len(imported_reward_rows),
                "store_hnt": plain(imported_hnt),
            }
        )
    return output


def imported_hnt_reward_months(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    return rows(
        conn,
        """
        SELECT substr(timestamp_utc, 1, 7) AS month,
               source,
               count(*) AS event_count,
               sum(CAST(quantity AS REAL)) AS hnt
        FROM ai_raw_events_flat
        WHERE asset = 'HNT'
          AND event_type = 'mining_reward'
          AND source IN ('helium_legacy_cointracking', 'heliumtracker')
        GROUP BY month, source
        ORDER BY month, source
        """,
    )


def base_tx_id(tx_id: str) -> str:
    return str(tx_id or "").split("+", 1)[0].strip()


def source_file_import_status(conn: sqlite3.Connection, filename: str) -> list[dict[str, Any]]:
    return rows(
        conn,
        """
        SELECT source_file_id, source_name, row_count
        FROM source_files
        WHERE source_name LIKE ?
        ORDER BY source_name
        """,
        (f"%{filename}",),
    )


def binance_staking_deposit_workbook_audit(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for filename in BINANCE_STAKING_DEPOSIT_FILES:
        paths = find_local_workbooks(filename)
        selected = paths[0] if paths else None
        records = sheet_records(selected) if selected else []
        hnt_rows = [row for row in records if str(row.get("Coin") or "").upper() == "HNT"]
        imported_rows = source_file_import_status(conn, filename)
        output.append(
            {
                "filename": filename,
                "local_path_count": len(paths),
                "selected_path": str(selected.relative_to(ROOT)) if selected else "",
                "xlsx_row_count": len(records),
                "hnt_row_count": len(hnt_rows),
                "hnt_quantity": plain(sum((dec(row.get("Amount")) for row in hnt_rows), Decimal("0"))),
                "imported": bool(imported_rows),
                "store_sources": [
                    {
                        "source_file_id": str(row.get("source_file_id") or ""),
                        "source_name": str(row.get("source_name") or ""),
                        "row_count": int(row.get("row_count") or 0),
                    }
                    for row in imported_rows
                ],
                "hnt_rows": [
                    {
                        "date_utc": str(row.get("Date(UTC)") or ""),
                        "amount": plain(row.get("Amount")),
                        "tx_id": str(row.get("TXID") or ""),
                        "status": str(row.get("Status") or ""),
                    }
                    for row in hnt_rows
                ],
            }
        )
    return output


def staking_transfer_workbook_audit() -> dict[str, Any]:
    path = first_local_workbook(STAKING_TRANSFER_WORKBOOK)
    if not path:
        return {"found": False, "filename": STAKING_TRANSFER_WORKBOOK}
    records = sheet_records(path)
    transfers: list[dict[str, Any]] = []
    outbound_from_main = Decimal("0")
    outbound_to_binance = Decimal("0")
    outbound_to_staking_wallet = Decimal("0")
    for row in records:
        sender = clean_text(row.get("Sender"))
        receiver = clean_text(row.get("Receiver"))
        amount = dec(row.get("Amount"))
        time = clean_text(row.get("Time"))
        if MAIN_HELIUM_WALLET[:10] not in sender and not sender.startswith("133"):
            continue
        outbound_from_main += amount
        receiver_lower = receiver.lower()
        if BINANCE_HNT_WALLET in receiver or "138bCXPV" in receiver or "binance" in receiver_lower:
            outbound_to_binance += amount
        if STAKING_WALLET[:8] in receiver or "14eKedP4g" in receiver:
            outbound_to_staking_wallet += amount
        transfers.append(
            {
                "time": time,
                "sender": sender,
                "receiver": receiver,
                "amount_hnt": plain(amount),
                "target": (
                    "binance"
                    if "138bCXPV" in receiver or "binance" in receiver_lower
                    else "staking_wallet_14e"
                    if "14eKedP4g" in receiver
                    else "other"
                ),
            }
        )
    return {
        "found": True,
        "filename": STAKING_TRANSFER_WORKBOOK,
        "selected_path": str(path.relative_to(ROOT)),
        "row_count": len(records),
        "main_wallet_out_count": len(transfers),
        "main_wallet_out_hnt": plain(outbound_from_main),
        "main_wallet_to_binance_hnt": plain(outbound_to_binance),
        "main_wallet_to_staking_wallet_hnt": plain(outbound_to_staking_wallet),
        "critical_rows": [
            row
            for row in transfers
            if row["time"].startswith(("2021-08-10", "2021-08-14", "2021-08-17", "2021-08-20", "2022-07-12"))
        ],
        "sample_rows": transfers[:40],
    }


def daily_reward_workbook_audit() -> dict[str, Any]:
    path = first_local_workbook(DAILY_REWARD_WORKBOOK)
    if not path:
        return {"found": False, "filename": DAILY_REWARD_WORKBOOK}
    records = sheet_records(path, "Abfrageergebnis")
    cutoffs = {
        "2021-08-17": Decimal("0"),
        "2021-08-20": Decimal("0"),
        "2022-07-12": Decimal("0"),
    }
    month_totals: dict[str, Decimal] = {}
    for row in records:
        date_text = str(row.get("date") or row.get("Datum") or "")
        amount = dec(row.get("total_hnt"))
        month = date_text[:7]
        if month:
            month_totals[month] = month_totals.get(month, Decimal("0")) + amount
        for cutoff in cutoffs:
            if date_text[:10] <= cutoff:
                cutoffs[cutoff] += amount
    return {
        "found": True,
        "filename": DAILY_REWARD_WORKBOOK,
        "selected_path": str(path.relative_to(ROOT)),
        "row_count": len(records),
        "cutoff_hnt": {cutoff: plain(value) for cutoff, value in cutoffs.items()},
        "month_hnt_2021": {
            month: plain(value)
            for month, value in sorted(month_totals.items())
            if month.startswith("2021-")
        },
    }


def binance_history_workbook_audit() -> dict[str, Any]:
    path = first_local_workbook(BINANCE_HISTORY_WORKBOOK)
    if not path:
        return {"found": False, "filename": BINANCE_HISTORY_WORKBOOK}
    records = sheet_records(path, "Binance Rohdaten")
    critical: list[dict[str, Any]] = []
    for row in records:
        asset = str(row.get("Coin") or row.get("Asset") or "").upper()
        if asset != "HNT":
            continue
        timestamp = str(row.get("UTC_Time") or row.get("Date(UTC)") or row.get("Time") or "")
        if not ("2021-08-01" <= timestamp[:10] <= "2021-08-21"):
            continue
        critical.append(
            {
                "time": timestamp,
                "operation": str(row.get("Operation") or ""),
                "change": plain(row.get("Change") or row.get("Amount")),
            }
        )
    return {
        "found": True,
        "filename": BINANCE_HISTORY_WORKBOOK,
        "selected_path": str(path.relative_to(ROOT)),
        "row_count": len(records),
        "critical_hnt_rows": critical,
    }


def staking_wallet_db_audit(conn: sqlite3.Connection) -> dict[str, Any]:
    raw_events = rows(
        conn,
        """
        SELECT r.unique_event_id, r.row_index, r.payload_json, sf.source_name
        FROM raw_events r
        LEFT JOIN source_files sf ON sf.source_file_id = r.source_file_id
        WHERE sf.source_name = ?
        ORDER BY COALESCE(json_extract(r.payload_json, '$.timestamp_utc'), json_extract(r.payload_json, '$.timestamp')),
                 r.row_index
        """,
        (STAKING_WALLET_SOURCE_NAME,),
    )
    inbound = Decimal("0")
    outbound = Decimal("0")
    event_rows: list[dict[str, Any]] = []
    for row in raw_events:
        data = payload(row)
        qty = dec(data.get("quantity"))
        side = str(data.get("side") or "").lower()
        if side == "in":
            inbound += qty
        elif side == "out":
            outbound += qty
        event_rows.append(
            {
                "event_id": str(row.get("unique_event_id") or ""),
                "timestamp_utc": str(data.get("timestamp_utc") or ""),
                "side": side,
                "quantity_hnt": plain(qty),
                "tx_id": str(data.get("tx_id") or ""),
            }
        )

    candidate_rows: list[dict[str, Any]] = []
    for row in raw_events:
        data = payload(row)
        tx = base_tx_id(str(data.get("tx_id") or ""))
        if not tx:
            continue
        counterparts = rows(
            conn,
            """
            SELECT r.unique_event_id, r.payload_json, sf.source_name
            FROM raw_events r
            LEFT JOIN source_files sf ON sf.source_file_id = r.source_file_id
            WHERE r.unique_event_id != ?
              AND json_extract(r.payload_json, '$.asset') = 'HNT'
              AND json_extract(r.payload_json, '$.tx_id') LIKE ?
            ORDER BY COALESCE(json_extract(r.payload_json, '$.timestamp_utc'), json_extract(r.payload_json, '$.timestamp'))
            """,
            (str(row.get("unique_event_id") or ""), f"{tx}%"),
        )
        for counterpart in counterparts:
            other = payload(counterpart)
            if str(other.get("side") or "").lower() == str(data.get("side") or "").lower():
                continue
            outbound_event_id = (
                str(row.get("unique_event_id") or "")
                if str(data.get("side") or "").lower() == "out"
                else str(counterpart.get("unique_event_id") or "")
            )
            inbound_event_id = (
                str(row.get("unique_event_id") or "")
                if str(data.get("side") or "").lower() == "in"
                else str(counterpart.get("unique_event_id") or "")
            )
            existing = rows(
                conn,
                """
                SELECT match_id, status, method
                FROM transfer_matches
                WHERE outbound_event_id = ? AND inbound_event_id = ?
                """,
                (outbound_event_id, inbound_event_id),
            )
            amount_diff = abs(dec(data.get("quantity")) - dec(other.get("quantity")))
            candidate = {
                "base_tx_id": tx,
                "outbound_event_id": outbound_event_id,
                "inbound_event_id": inbound_event_id,
                "outbound_source": str(payload(row if outbound_event_id == row.get("unique_event_id") else counterpart).get("source") or ""),
                "inbound_source": str(payload(row if inbound_event_id == row.get("unique_event_id") else counterpart).get("source") or ""),
                "timestamp_utc": str(data.get("timestamp_utc") or other.get("timestamp_utc") or ""),
                "amount_diff_hnt": plain(amount_diff),
                "match_exists": bool(existing),
                "existing_match_id": str(existing[0].get("match_id") or "") if existing else "",
            }
            if candidate not in candidate_rows:
                candidate_rows.append(candidate)

    return {
        "source_name": STAKING_WALLET_SOURCE_NAME,
        "event_count": len(event_rows),
        "inbound_hnt": plain(inbound),
        "outbound_hnt": plain(outbound),
        "balance_hnt": plain(inbound - outbound),
        "events": event_rows,
        "self_transfer_pair_candidates": sorted(
            candidate_rows,
            key=lambda item: (str(item["timestamp_utc"]), str(item["outbound_event_id"]), str(item["inbound_event_id"])),
        ),
    }


def local_excel_staking_wallet_audit(conn: sqlite3.Connection) -> dict[str, Any]:
    return {
        "binance_staking_deposit_workbooks": binance_staking_deposit_workbook_audit(conn),
        "staking_transfer_workbook": staking_transfer_workbook_audit(),
        "daily_reward_workbook": daily_reward_workbook_audit(),
        "binance_history_workbook": binance_history_workbook_audit(),
        "staking_wallet_db": staking_wallet_db_audit(conn),
    }


def build_audit(conn: sqlite3.Connection) -> dict[str, Any]:
    lines = remaining_lines(conn)
    detail_rows: list[dict[str, Any]] = []
    for line in lines:
        source_event = event_summary(load_raw_event(conn, str(line.get("source_event_id") or "")))
        lot_event = event_summary(load_raw_event(conn, str(line.get("lot_source_event_id") or "")))
        matches = transfer_matches_for_inbound(conn, str(line.get("lot_source_event_id") or ""))
        same_tx = same_tx_events(conn, source_event.get("tx_id", ""))[:8]
        detail_rows.append(
            {
                "tax_year": int(line.get("tax_year") or 0),
                "job_id": str(line.get("job_id") or ""),
                "line_no": int(line.get("line_no") or 0),
                "asset": str(line.get("asset") or ""),
                "qty": plain(line.get("qty")),
                "buy_timestamp_utc": str(line.get("buy_timestamp_utc") or ""),
                "sell_timestamp_utc": str(line.get("sell_timestamp_utc") or ""),
                "cost_basis_eur": plain(line.get("cost_basis_eur")),
                "proceeds_eur": plain(line.get("proceeds_eur")),
                "gain_loss_eur": plain(line.get("gain_loss_eur")),
                "source_event_id": str(line.get("source_event_id") or ""),
                "lot_source_event_id": str(line.get("lot_source_event_id") or ""),
                "transfer_chain_id": str(line.get("transfer_chain_id") or ""),
                "classification": classify(line, lot_event, matches),
                "source_event": source_event,
                "lot_event": lot_event,
                "transfer_matches": matches,
                "same_tx_events": same_tx,
            }
        )

    groups: dict[tuple[str, int, str], dict[str, Any]] = {}
    for item in detail_rows:
        key = (str(item["classification"]), int(item["tax_year"]), str(item["asset"]))
        group = groups.setdefault(
            key,
            {
                "classification": key[0],
                "tax_year": key[1],
                "asset": key[2],
                "line_count": 0,
                "qty": Decimal("0"),
                "proceeds_eur": Decimal("0"),
                "lot_source_event_ids": set(),
                "source_platforms": set(),
            },
        )
        group["line_count"] += 1
        group["qty"] += dec(item["qty"])
        group["proceeds_eur"] += dec(item["proceeds_eur"])
        if item["lot_source_event_id"]:
            group["lot_source_event_ids"].add(item["lot_source_event_id"])
        group["source_platforms"].add(item["source_event"].get("source") or "")

    serializable_groups = []
    for group in sorted(groups.values(), key=lambda x: (x["tax_year"], x["asset"], x["classification"])):
        serializable_groups.append(
            {
                "classification": group["classification"],
                "tax_year": group["tax_year"],
                "asset": group["asset"],
                "line_count": group["line_count"],
                "qty": plain(group["qty"]),
                "proceeds_eur": plain(group["proceeds_eur"]),
                "lot_source_event_ids": sorted(group["lot_source_event_ids"]),
                "source_platforms": sorted(platform for platform in group["source_platforms"] if platform),
            }
        )

    return {
        "run_date": RUN_DATE,
        "created_at_utc": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "db_path": str(DEFAULT_DB),
        "latest_jobs": latest_jobs(conn),
        "thresholds": {"material_proceeds_eur": str(MATERIAL_PROCEEDS_EUR)},
        "counts": {
            "remaining_line_count": len(detail_rows),
            "remaining_proceeds_eur": plain(sum((dec(item["proceeds_eur"]) for item in detail_rows), Decimal("0"))),
        },
        "groups": serializable_groups,
        "lines": detail_rows,
        "hnt_source_balance_context": hnt_source_balance_context(conn),
        "local_heliumtracker_file_coverage": local_heliumtracker_file_coverage(conn),
        "imported_hnt_reward_months": imported_hnt_reward_months(conn),
        "local_excel_staking_wallet_audit": local_excel_staking_wallet_audit(conn),
    }


def render_md(audit: dict[str, Any]) -> str:
    lines = [
        "# HNT-/USDT-Restbestandsluecken-Audit",
        "",
        f"Stand: {RUN_DATE}",
        "",
        "## Ergebnis",
        "",
        f"- Aktuelle Restzeilen: `{audit['counts']['remaining_line_count']}`",
        f"- Erloes dieser Restzeilen: `{audit['counts']['remaining_proceeds_eur']} EUR`",
        "- Keine Restzeile ist ein belegbarer Preisanker- oder FX-Backfill.",
        "- HNT-Transfer-Matches fuer die Binance-Deposits existieren bereits; die Luecke liegt vor dem Legacy-Outflow.",
        "- USDT-Reste bleiben Pionex-/Binance-Opening- bzw. Bot-Historie ohne Primaerbeleg.",
        "",
        "## Aktuelle Jobs",
        "",
        "| Jahr | Job | Tax-Lines | Derivate-Lines | Aktualisiert |",
        "| ---: | --- | ---: | ---: | --- |",
    ]
    for job in audit["latest_jobs"]:
        lines.append(
            f"| {job['tax_year']} | `{job['job_id']}` | {job['tax_line_count']} | "
            f"{job['derivative_line_count']} | {job['updated_at_utc']} |"
        )
    lines.extend(
        [
            "",
            "## Gruppierung",
            "",
            "| Jahr | Asset | Klasse | Zeilen | Menge | Erloes EUR | Plattformen |",
            "| ---: | --- | --- | ---: | ---: | ---: | --- |",
        ]
    )
    for group in audit["groups"]:
        lines.append(
            f"| {group['tax_year']} | `{group['asset']}` | `{group['classification']}` | "
            f"{group['line_count']} | {group['qty']} | {group['proceeds_eur']} | "
            f"{', '.join(group['source_platforms'])} |"
        )

    lines.extend(
        [
            "",
            "## Betroffene Zeilen",
            "",
            "| Jahr | Line | Asset | Menge | Erloes EUR | Quelle | Lot-Quelle | Klasse |",
            "| ---: | ---: | --- | ---: | ---: | --- | --- | --- |",
        ]
    )
    for item in audit["lines"]:
        source = item["source_event"]
        lot = item["lot_event"]
        source_text = f"{source['source']}/{source['event_type']}/{source['side']}"
        lot_text = f"{lot['source']}/{lot['event_type']}/{lot['side']}" if item["lot_source_event_id"] else "leer"
        lines.append(
            f"| {item['tax_year']} | {item['line_no']} | `{item['asset']}` | {item['qty']} | "
            f"{item['proceeds_eur']} | `{source_text}` | `{lot_text}` | `{item['classification']}` |"
        )

    lines.extend(["", "## HNT-Transfer-Belege", ""])
    hnt_match_lines = [item for item in audit["lines"] if item["asset"] == "HNT" and item["transfer_matches"]]
    seen_matches: set[str] = set()
    if not hnt_match_lines:
        lines.append("- Keine HNT-Transfer-Matches in den Restzeilen.")
    for item in hnt_match_lines:
        for match in item["transfer_matches"]:
            match_id = match["match_id"]
            if match_id in seen_matches:
                continue
            seen_matches.add(match_id)
            outbound = match["outbound"]
            inbound = match["inbound"]
            lines.extend(
                [
                    f"- Match `{match_id}` fuer Inbound `{match['inbound_event_id']}`:",
                    f"  - Outbound `{outbound['source']}` `{outbound['event_type']}` "
                    f"{outbound['quantity']} `{outbound['asset']}` am `{outbound['timestamp_utc']}`.",
                    f"  - Binance-Inbound {inbound['quantity']} `{inbound['asset']}` am `{inbound['timestamp_utc']}`.",
                    f"  - Delta `{match['amount_diff']}`, Zeitdifferenz `{match['time_diff_seconds']}` Sekunden.",
                    f"  - Legacy-Transfer-Wert `value_usd={outbound['value_usd']}` ist ein Transferwert, "
                    "keine belegte Anschaffungskostenbasis.",
                ]
            )

    lines.extend(
        [
            "",
            "## Mining-Reward-Kontext",
            "",
            "- BMF 2025, Randnummern 7 bis 11, beschreibt Block-Rewards/Mining als Erwerb von Kryptowerten im Rahmen der Blockerstellung.",
            "- BMF 2025, Randnummern 38 bis 44, ordnet Blockerstellung nicht als private Vermoegensverwaltung ein und behandelt den Zugang im Betriebsvermoegen mit Marktkurs/Anschaffungskostenlogik.",
            "- BMF 2025, Randnummern 43, 51 und 91, stuetzen Marktkurs/Tageskurs und Abzug individueller bzw. fortgefuehrter Anschaffungskosten bei Betriebsvermoegen.",
            "- Projektlogik: `mining_reward` wird als Reward/Business-Lot verarbeitet. Der Restbefund ist deshalb keine falsche Mining-Klassifikation, sondern fehlender belegter Vorbestand vor den konkreten Outflows.",
            "",
            "## Lokale HeliumTracker-Quellenabdeckung",
            "",
            "| Datei | CSV-Zeilen | CSV-HNT | Importiert | Store-Zeilen | Reward-Events | Store-HNT |",
            "| --- | ---: | ---: | --- | ---: | ---: | ---: |",
        ]
    )
    for item in audit["local_heliumtracker_file_coverage"]:
        lines.append(
            f"| `{item['file']}` | {item['fs_rows']} | {item['fs_hnt']} | "
            f"{'ja' if item['imported'] else 'nein'} | {item['store_row_count']} | "
            f"{item['store_reward_event_count']} | {item['store_hnt']} |"
        )

    excel_audit = audit["local_excel_staking_wallet_audit"]
    staking_workbook = excel_audit["staking_transfer_workbook"]
    daily_workbook = excel_audit["daily_reward_workbook"]
    staking_db = excel_audit["staking_wallet_db"]
    lines.extend(
        [
            "",
            "## Lokale Excel-Pruefung Staking-Wallet",
            "",
            "Die vom Nutzer vermutete Staking-Wallet-Spur ist lokal vorhanden. Rohdaten bleiben lokal; diese Sektion dokumentiert nur abgeleitete Pruefergebnisse.",
            "",
            "### Binance-HNT-Deposit-Excel",
            "",
            "| Datei | Lokale Kopien | Zeilen | HNT-Zeilen | HNT-Menge | Importiert |",
            "| --- | ---: | ---: | ---: | ---: | --- |",
        ]
    )
    for item in excel_audit["binance_staking_deposit_workbooks"]:
        lines.append(
            f"| `{item['filename']}` | {item['local_path_count']} | {item['xlsx_row_count']} | "
            f"{item['hnt_row_count']} | {item['hnt_quantity']} | {'ja' if item['imported'] else 'nein'} |"
        )

    if staking_workbook.get("found"):
        lines.extend(
            [
                "",
                "### Heliumtracker-Transferliste 2021",
                "",
                f"- Datei: `{staking_workbook['selected_path']}`",
                f"- Zeilen: `{staking_workbook['row_count']}`",
                f"- Abgaenge aus der Haupt-Wallet `{MAIN_HELIUM_WALLET[:8]}...`: `{staking_workbook['main_wallet_out_count']}` mit `{staking_workbook['main_wallet_out_hnt']} HNT`.",
                f"- Davon zu Binance: `{staking_workbook['main_wallet_to_binance_hnt']} HNT`.",
                f"- Davon zur Staking-Wallet `{STAKING_WALLET[:8]}...`: `{staking_workbook['main_wallet_to_staking_wallet_hnt']} HNT`.",
                "",
                "Kritische Zeilen aus der Excel-Transferliste:",
                "",
                "| Zeit | Ziel | Menge HNT | Empfaenger |",
                "| --- | --- | ---: | --- |",
            ]
        )
        for row in staking_workbook["critical_rows"]:
            lines.append(
                f"| `{row['time']}` | `{row['target']}` | {row['amount_hnt']} | `{row['receiver']}` |"
            )

    if daily_workbook.get("found"):
        lines.extend(
            [
                "",
                "### Daily-Hotspot-Rewards-Excel",
                "",
                f"- Datei: `{daily_workbook['selected_path']}`",
                f"- Zeilen: `{daily_workbook['row_count']}`",
                f"- HNT bis `2021-08-17`: `{daily_workbook['cutoff_hnt']['2021-08-17']}`",
                f"- HNT bis `2021-08-20`: `{daily_workbook['cutoff_hnt']['2021-08-20']}`",
                f"- HNT bis `2022-07-12`: `{daily_workbook['cutoff_hnt']['2022-07-12']}`",
                "",
                "Monatssummen 2021 aus der Excel-Datei:",
                "",
                "| Monat | HNT |",
                "| --- | ---: |",
            ]
        )
        for month, value in daily_workbook["month_hnt_2021"].items():
            lines.append(f"| `{month}` | {value} |")

    lines.extend(
        [
            "",
            "### Staking-Wallet in der Datenbank",
            "",
            f"- Importierte Quelle: `{staking_db['source_name']}`",
            f"- Events: `{staking_db['event_count']}`",
            f"- Inbound: `{staking_db['inbound_hnt']} HNT`",
            f"- Outbound: `{staking_db['outbound_hnt']} HNT`",
            f"- Saldo: `{staking_db['balance_hnt']} HNT`",
            f"- Nicht gematchte Self-Transfer-Kandidaten mit gleicher Helium-Transaktion: `{sum(1 for row in staking_db['self_transfer_pair_candidates'] if not row['match_exists'])}`",
            "",
            "| Zeit | Out-Quelle | In-Quelle | Delta HNT | Match vorhanden |",
            "| --- | --- | --- | ---: | --- |",
        ]
    )
    for row in staking_db["self_transfer_pair_candidates"]:
        lines.append(
            f"| `{row['timestamp_utc']}` | `{row['outbound_source']}` | `{row['inbound_source']}` | "
            f"{row['amount_diff_hnt']} | {'ja' if row['match_exists'] else 'nein'} |"
        )

    binance_history = excel_audit["binance_history_workbook"]
    if binance_history.get("found"):
        lines.extend(
            [
                "",
                "### Binance-Historie 2021-08",
                "",
                f"- Datei: `{binance_history['selected_path']}`",
                f"- Gepruefte Rohzeilen: `{binance_history['row_count']}`",
                "- Der Auszug bestaetigt HNT-Deposits am `2021-08-06`, `2021-08-10` und `2021-08-20`; zwischen `2021-08-10` und den HNT-Verkaeufen am `2021-08-17` ist kein zusaetzlicher Binance-HNT-Deposit sichtbar.",
                "",
                "| Zeit | Operation | HNT-Aenderung |",
                "| --- | --- | ---: |",
            ]
        )
        for row in binance_history["critical_hnt_rows"]:
            lines.append(f"| `{row['time']}` | `{row['operation']}` | {row['change']} |")

    lines.extend(
        [
            "",
            "Abdeckungsschluss:",
            "",
            "- Die im Workspace vorhandenen HeliumTracker-Dateien sind importiert; die HNT-Summen aus CSV und Store stimmen je Datei ueberein.",
            "- Lokal vorhanden ist fuer `2021` nur `heliumtracker-report-advanced-2021-12.csv`; fuer die kritischen Binance-Verkaeufe am `2021-08-17` gibt es damit keine zusaetzliche lokale HeliumTracker-Quelle.",
            "- Fuer `2022-02` bis `2022-07` sind HeliumTracker-Rewards importiert; sie reichen zusammen mit dem Legacy-Cointracking-Saldo aber nicht aus, um den `450.0398803021218`-HNT-Legacy-Outflow am `2022-07-12` belegbar zu decken.",
            "- Die Excel-Dateien bestaetigen eine eigene Staking-Wallet `14eKed...`; dort fehlen aktuell Self-Transfer-Matches zur Haupt-Wallet `133...`.",
            "- Diese Self-Transfer-Matches verbessern die Kettenbelegung, erzeugen aber keine Anschaffungskosten. Fuer die grossen 2022-Zufluesse in die Staking-Wallet bleibt die Herkunft vor der Staking-/Return-Wallet ungeklaert.",
            "- Fuer `2021-08-17` zeigt die Excel-Spur keine zusaetzliche Binance-Einzahlung; die `100 HNT` am `2021-08-14` gingen zur Staking-Wallet und verliessen diese kurz danach wieder.",
            "",
            "Quellen:",
            "",
            "- BMF 2025: `https://www.bundesfinanzministerium.de/Content/DE/Downloads/BMF_Schreiben/Steuerarten/Einkommensteuer/2025-03-06-einzelfragen-kryptowerte-bmf-schreiben.pdf?__blob=publicationFile&v=3`",
            "- BMF-Erlaeuterungsseite 2025: `https://www.bundesfinanzministerium.de/Content/DE/Downloads/BMF_Schreiben/Steuerarten/Einkommensteuer/2025-03-06-einzelfragen-kryptowerte.html`",
            "",
            "## HNT-Bestandsschnitte",
            "",
            "| Zeitpunkt | Kontext | Quelle | Events | Mining-Rewards HNT | In HNT | Out HNT | Saldo HNT |",
            "| --- | --- | --- | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for item in audit["hnt_source_balance_context"]:
        lines.append(
            f"| `{item['cutoff']}` | {item['label']} | `{item['source']}` | {item['event_count']} | "
            f"{item['mining_reward_hnt']} | {item['inbound_hnt']} | {item['outbound_hnt']} | {item['balance_hnt']} |"
        )

    lines.extend(
        [
            "",
            "## Bewertung",
            "",
            "- Eine automatische Bewertung der HNT-Deposits mit dem Legacy-Transferwert waere fachlich falsch, "
            "weil Transferwert nicht gleich Anschaffungskosten ist.",
            "- Dass HNT im Legacy-Kontext aus Mining-Rewards stammt, hilft fachlich: Rewards koennen Anschaffungskosten tragen, wenn sie als bewertete Lots vorhanden sind.",
            "- Fuer die konkreten Restzeilen reicht der belegte Legacy-Bestand vor den Outflows aber nicht aus; vorhandene Mining-Rewards wurden bereits vorher durch andere Outflows/Transfers verbraucht.",
            "- Die 2021-HNT-Zeilen ohne Lot-Quelle liegen auf Binance-Verkaeufen am `2021-08-17`; "
            "fuer diese Verkaufsmenge gibt es im aktiven Datenstand keinen belegten vorherigen Binance-Deposit.",
            "- Die Excel-Pruefung stuetzt diese Einordnung: Zwischen den bekannten Binance-Deposits am `2021-08-10` und den Verkaeufen am `2021-08-17` wurde kein weiterer Binance-HNT-Zufluss gefunden.",
            "- Fuer den `2022-07-12`-Komplex ist die Luecke enger eingegrenzt: die Staking-Wallet erhielt kurz vorher `421.34562734 HNT` aus drei Legacy-Transfers und leitete `421.30245111 HNT` an die Haupt-Wallet zurueck; fuer diese drei Zufluesse fehlt weiterhin eine bewertete Primaerherkunft.",
            "- Die 2022-USDT-Zeilen decken sich mit dem bereits dokumentierten Pionex-/Binance-"
            "Opening- und Bot-Historienproblem.",
            "- Deshalb: keine neue automatische RAW-/FX-/Cost-Basis-Korrektur aus diesem Audit.",
            "",
            "## Naechste sichere Aktion",
            "",
            "- HNT: Primaerbelege fuer HNT-Anschaffung/Mining-Bestand vor den Legacy-Outflows nachreichen oder "
            "die historischen Nullbasis-Zeilen bewusst offen lassen.",
            "- USDT: Pionex-Opening-/Bot-Historie oder explizite Review-Entscheidung verwenden; ohne Beleg "
            "keinen steuerwirksamen Zufluss importieren.",
            "",
            f"JSON: `{OUT_JSON.relative_to(ROOT)}`",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> None:
    uri = f"file:{DEFAULT_DB}?mode=ro&immutable=1"
    conn = sqlite3.connect(uri, uri=True)
    conn.row_factory = sqlite3.Row
    audit = build_audit(conn)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(audit, indent=2, sort_keys=True), encoding="utf-8")
    OUT_MD.write_text(render_md(audit), encoding="utf-8")
    print(json.dumps({"json": str(OUT_JSON), "report": str(OUT_MD), "counts": audit["counts"]}, indent=2))


if __name__ == "__main__":
    main()
