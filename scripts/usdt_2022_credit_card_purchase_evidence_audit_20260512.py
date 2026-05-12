#!/usr/bin/env python3
"""Audit local evidence for a possible 2022 USDT credit-card purchase."""

from __future__ import annotations

import json
import sqlite3
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RUN_DATE = "2026-05-12"
DEFAULT_DB = Path("/root/.local/share/steuerreport/ai_readonly/steuerreport_ai_readonly.sqlite")
OUT_JSON = ROOT / "var" / "usdt_2022_credit_card_purchase_evidence_audit_2026-05-12.json"
OUT_MD = ROOT / "docs" / "239_USDT_2022_CREDIT_CARD_PURCHASE_EVIDENCE_AUDIT_2026-05-12.md"

LOCAL_FIAT_FILES = (
    ROOT / "usertransfer" / "Binance" / "export 2021" / "Binance-Fiat-Buy-History-202605061831(UTC+2)_c87830e5.xlsx",
    ROOT / "usertransfer" / "Binance" / "export 2021" / "Binance-Fiat-Deposit-History-202605061832(UTC+2)_9311aa32.xlsx",
    ROOT / "usertransfer" / "Binance" / "export 2021" / "Binance-Fiat-Withdraw-History-202605061833(UTC+2)_2416f4a2.xlsx",
)


def dec(value: Any) -> Decimal:
    try:
        return Decimal(str(value or "0").strip().replace(",", "."))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def plain(value: Any) -> str:
    value_dec = dec(value)
    text = format(value_dec, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def rows(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    cur = conn.execute(sql, params)
    return [dict(row) for row in cur.fetchall()]


def load_json(text: str) -> dict[str, Any]:
    try:
        loaded = json.loads(text or "{}")
    except json.JSONDecodeError:
        return {}
    return loaded if isinstance(loaded, dict) else {}


def text(value: Any) -> str:
    return "" if value is None else str(value)


def workbook_summary(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"path": str(path.relative_to(ROOT)), "exists": False, "data_rows": 0, "sheets": []}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        data_rows = 0
        for sheet in workbook.worksheets:
            nonempty_rows = 0
            preview: list[list[str]] = []
            for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = ["" if cell is None else str(cell) for cell in row]
                if any(cell.strip() for cell in cells):
                    nonempty_rows += 1
                    if len(preview) < 3:
                        preview.append(cells)
                if index >= 1000 and nonempty_rows == 0:
                    break
            data_rows += max(nonempty_rows - 1, 0)
            sheets.append(
                {
                    "title": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "nonempty_rows": nonempty_rows,
                    "data_rows_after_header": max(nonempty_rows - 1, 0),
                    "preview": preview,
                }
            )
        return {"path": str(path.relative_to(ROOT)), "exists": True, "data_rows": data_rows, "sheets": sheets}
    finally:
        workbook.close()


def latest_remaining_usdt(conn: sqlite3.Connection) -> dict[str, str]:
    lines = rows(
        conn,
        """
        SELECT qty, proceeds_eur
        FROM ai_open_zero_cost_tax_lines
        WHERE tax_year = 2022
          AND asset = 'USDT'
          AND CAST(proceeds_eur AS REAL) >= 50
        """,
    )
    return {
        "line_count": str(len(lines)),
        "qty_usdt": plain(sum((dec(row["qty"]) for row in lines), Decimal("0"))),
        "proceeds_eur": plain(sum((dec(row["proceeds_eur"]) for row in lines), Decimal("0"))),
    }


def fiat_purchase_events(conn: sqlite3.Connection) -> list[dict[str, str]]:
    output = []
    for row in rows(
        conn,
        """
        SELECT
            json_extract(re.payload_json, '$.timestamp_utc') AS timestamp_utc,
            json_extract(re.payload_json, '$.asset') AS asset,
            json_extract(re.payload_json, '$.side') AS side,
            json_extract(re.payload_json, '$.quantity') AS quantity,
            json_extract(re.payload_json, '$.event_type') AS event_type,
            sf.source_name,
            re.row_index,
            re.unique_event_id
        FROM raw_events re
        JOIN source_files sf ON sf.source_file_id = re.source_file_id
        WHERE json_extract(re.payload_json, '$.event_type') = 'fiat_crypto_purchase'
        ORDER BY timestamp_utc, asset, side
        """,
    ):
        output.append({key: text(value) for key, value in row.items()})
    return output


def binance_dec_jan_fiat_usdt_events(conn: sqlite3.Connection) -> list[dict[str, str]]:
    output = []
    for row in rows(
        conn,
        """
        SELECT
            json_extract(re.payload_json, '$.timestamp_utc') AS timestamp_utc,
            json_extract(re.payload_json, '$.asset') AS asset,
            json_extract(re.payload_json, '$.side') AS side,
            json_extract(re.payload_json, '$.event_type') AS event_type,
            json_extract(re.payload_json, '$.quantity') AS quantity,
            sf.source_name,
            re.row_index,
            re.unique_event_id
        FROM raw_events re
        JOIN source_files sf ON sf.source_file_id = re.source_file_id
        WHERE json_extract(re.payload_json, '$.source') LIKE 'binance%'
          AND json_extract(re.payload_json, '$.asset') IN ('USDT', 'EUR')
          AND json_extract(re.payload_json, '$.timestamp_utc') BETWEEN
              '2021-12-01T00:00:00+00:00' AND '2022-01-31T23:59:59+00:00'
        ORDER BY timestamp_utc, asset, side, row_index
        """,
    ):
        output.append({key: text(value) for key, value in row.items()})
    return output


def fiat_source_coverage(conn: sqlite3.Connection) -> list[dict[str, str]]:
    output = []
    for row in rows(
        conn,
        """
        SELECT
            sf.source_name,
            count(*) AS event_count,
            min(json_extract(re.payload_json, '$.timestamp_utc')) AS min_timestamp_utc,
            max(json_extract(re.payload_json, '$.timestamp_utc')) AS max_timestamp_utc
        FROM raw_events re
        JOIN source_files sf ON sf.source_file_id = re.source_file_id
        WHERE lower(sf.source_name) LIKE '%fiat%'
           OR lower(sf.source_name) LIKE '%card%'
           OR json_extract(re.payload_json, '$.event_type') = 'fiat_crypto_purchase'
        GROUP BY sf.source_name
        ORDER BY sf.source_name
        """,
    ):
        output.append({key: text(value) for key, value in row.items()})
    return output


def search_card_terms(conn: sqlite3.Connection) -> dict[str, int]:
    terms = ("card", "credit", "kredit", "visa", "mastercard", "fiat_crypto_purchase")
    result: dict[str, int] = {}
    for term in terms:
        result[term] = int(
            rows(
                conn,
                """
                SELECT count(*) AS count
                FROM raw_events re
                JOIN source_files sf ON sf.source_file_id = re.source_file_id
                WHERE lower(re.payload_json) LIKE ?
                   OR lower(sf.source_name) LIKE ?
                """,
                (f"%{term.lower()}%", f"%{term.lower()}%"),
            )[0]["count"]
        )
    return result


def summarise_dec_jan(events: list[dict[str, str]]) -> dict[str, str]:
    totals: dict[tuple[str, str, str], Decimal] = {}
    for event in events:
        month = (event["timestamp_utc"] or "")[:7]
        key = (month, event["asset"], event["side"])
        totals[key] = totals.get(key, Decimal("0")) + dec(event["quantity"])
    return {
        "|".join(key): plain(value)
        for key, value in sorted(totals.items())
        if value
    }


def build_audit(db_path: Path = DEFAULT_DB) -> dict[str, Any]:
    conn = sqlite3.connect(f"file:{db_path}?mode=ro&immutable=1", uri=True)
    conn.row_factory = sqlite3.Row
    try:
        fiat_events = fiat_purchase_events(conn)
        dec_jan_events = binance_dec_jan_fiat_usdt_events(conn)
        jan_2022_fiat_purchase_events = [
            event
            for event in fiat_events
            if (event["timestamp_utc"] or "").startswith(("2021-12", "2022-01"))
        ]
        return {
            "created_at_utc": datetime.now(UTC).isoformat(),
            "db_path": str(db_path),
            "remaining_usdt": latest_remaining_usdt(conn),
            "fiat_purchase_events": fiat_events,
            "dec_jan_fiat_purchase_events": jan_2022_fiat_purchase_events,
            "binance_dec_jan_usdt_eur_summary": summarise_dec_jan(dec_jan_events),
            "binance_dec_jan_usdt_eur_event_count": len(dec_jan_events),
            "fiat_source_coverage": fiat_source_coverage(conn),
            "card_term_counts": search_card_terms(conn),
            "local_fiat_workbooks": [workbook_summary(path) for path in LOCAL_FIAT_FILES],
            "conclusion": {
                "status": "primary_evidence_missing",
                "finding": (
                    "Die Nutzererinnerung an einen damaligen USDT-Kauf per Kreditkarte ist plausibel, "
                    "aber lokal derzeit nicht als belastbarer Dec-2021/Jan-2022-Beleg importiert."
                ),
                "no_auto_fix_reason": (
                    "Ohne Binance-Fiat-Order-/Buy-Crypto-History, Kartenabrechnung oder aehnlichen "
                    "Primaerbeleg darf keine Anschaffungskostenbasis fuer die 2022-USDT-Reste erzeugt werden."
                ),
                "needed_evidence": (
                    "Binance Buy Crypto / Fiat Order History oder Kreditkarten-/Bankbeleg fuer USDT-Kaeufe "
                    "im Zeitraum 2021-12 bis 2022-01, inklusive Zeit, EUR-Betrag, USDT-Menge und Gebuehren."
                ),
            },
        }
    finally:
        conn.close()


def table(headers: list[str], body: list[list[str]]) -> list[str]:
    return [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
        *["| " + " | ".join(str(value) for value in row) + " |" for row in body],
    ]


def render_doc(audit: dict[str, Any]) -> str:
    remaining = audit["remaining_usdt"]
    fiat_events = audit["fiat_purchase_events"]
    dec_jan = audit["dec_jan_fiat_purchase_events"]
    workbooks = audit["local_fiat_workbooks"]
    coverage = audit["fiat_source_coverage"]
    terms = audit["card_term_counts"]

    total_usdt_fiat_in = sum(
        (
            dec(event["quantity"])
            for event in fiat_events
            if event["asset"] == "USDT" and event["side"] == "in"
        ),
        Decimal("0"),
    )
    total_eur_fiat_out = sum(
        (
            dec(event["quantity"])
            for event in fiat_events
            if event["asset"] == "EUR" and event["side"] == "out"
        ),
        Decimal("0"),
    )

    lines = [
        "# USDT 2022 Kreditkartenkauf-Belegaudit",
        "",
        f"Erstellt: `{audit['created_at_utc']}`",
        "",
        "## Anlass",
        "",
        "Der Nutzer erinnert, dass USDT damals wahrscheinlich per Kreditkarte gekauft wurden. "
        "Dieser Audit prueft, ob dafuer lokal bereits ein belastbarer Import- oder Primaerbeleg "
        "fuer den verbleibenden 2022-USDT-Block vorhanden ist.",
        "",
        "## Aktiver Restblock",
        "",
        f"- Offene 2022-USDT-Zeilen >= 50 EUR: `{remaining['line_count']}`",
        f"- Offene Menge: `{remaining['qty_usdt']} USDT`",
        f"- Betroffener Erloes: `{remaining['proceeds_eur']} EUR`",
        "",
        "## Ergebnis",
        "",
        "- Die Erinnerung an einen Kreditkarten-/Fiat-Kauf ist als Suchhinweis plausibel.",
        "- In der lokalen Readonly-Datenbank sind `fiat_crypto_purchase`-Events nur fuer `2021-02` bis `2021-04` vorhanden.",
        "- Fuer `2021-12` oder `2022-01` gibt es lokal kein `fiat_crypto_purchase`-Event.",
        "- Die lokale Binance-`Fiat-Buy-History`-XLSX ist vorhanden, enthaelt aber keine Datenzeilen.",
        "- Damit gibt es derzeit keinen deterministischen automatischen Cost-Basis-Fix fuer die drei 2022-USDT-Restzeilen.",
        "",
        "## Gefundene Fiat-Crypto-Kaeufe",
        "",
        f"- Anzahl Events: `{len(fiat_events)}`",
        f"- USDT-In aus Fiat-Crypto-Kauf: `{plain(total_usdt_fiat_in)} USDT`",
        f"- EUR-Out aus Fiat-Crypto-Kauf: `{plain(total_eur_fiat_out)} EUR`",
        "",
        *table(
            ["Zeit UTC", "Asset", "Seite", "Menge", "Quelle", "Row"],
            [
                [
                    event["timestamp_utc"],
                    event["asset"],
                    event["side"],
                    plain(event["quantity"]),
                    event["source_name"],
                    event["row_index"],
                ]
                for event in fiat_events
            ],
        ),
        "",
        "## Dec-2021/Jan-2022 Fiat-Kauf-Treffer",
        "",
    ]
    if dec_jan:
        lines.extend(
            table(
                ["Zeit UTC", "Asset", "Seite", "Menge", "Quelle", "Row"],
                [
                    [
                        event["timestamp_utc"],
                        event["asset"],
                        event["side"],
                        plain(event["quantity"]),
                        event["source_name"],
                        event["row_index"],
                    ]
                    for event in dec_jan
                ],
            )
        )
    else:
        lines.append("- Keine `fiat_crypto_purchase`-Events in `2021-12` oder `2022-01` gefunden.")

    lines.extend(
        [
            "",
            "## Binance-USDT/EUR im relevanten Fenster",
            "",
            f"- Binance-USDT/EUR-Events von `2021-12-01` bis `2022-01-31`: `{audit['binance_dec_jan_usdt_eur_event_count']}`",
            "- Diese Events zeigen vor allem Trade-/Withdrawal-/Fee-Bewegungen, aber keinen importierten Kreditkartenkauf als Anschaffungskette.",
            "",
            *table(
                ["Monat", "Asset", "Seite", "Summe"],
                [
                    [key.split("|")[0], key.split("|")[1], key.split("|")[2], value]
                    for key, value in audit["binance_dec_jan_usdt_eur_summary"].items()
                ],
            ),
            "",
            "## Lokale Fiat-Dateien",
            "",
            *table(
                ["Datei", "Vorhanden", "Datenzeilen", "Sheets"],
                [
                    [
                        item["path"],
                        str(item["exists"]),
                        str(item["data_rows"]),
                        ", ".join(
                            f"{sheet['title']}:{sheet['nonempty_rows']} nonempty"
                            for sheet in item.get("sheets", [])
                        ),
                    ]
                    for item in workbooks
                ],
            ),
            "",
            "## Source-Coverage",
            "",
            *table(
                ["Quelle", "Events", "Von", "Bis"],
                [
                    [
                        item["source_name"],
                        item["event_count"],
                        item["min_timestamp_utc"],
                        item["max_timestamp_utc"],
                    ]
                    for item in coverage
                ],
            ),
            "",
            "## Suchbegriffe",
            "",
            *table(
                ["Suchbegriff", "Treffer"],
                [[term, str(count)] for term, count in terms.items()],
            ),
            "",
            "## Naechste sichere Aktion",
            "",
            "Benoetigt wird ein Primaerbeleg fuer den vermuteten Kauf: Binance `Buy Crypto` / "
            "`Fiat Order History` / Kartenkauf-Historie oder eine Kreditkarten-/Bankabrechnung "
            "fuer `2021-12` bis `2022-01` mit Zeit, EUR-Betrag, USDT-Menge und Gebuehren. "
            "Erst dann kann ein Import-/Review-Fix fuer die Anschaffungskosten sauber erstellt werden.",
            "",
            "Kein automatischer Fix wurde abgeleitet, weil Nutzererinnerung allein keine Cost Basis, "
            "keinen FX-Kurs und keine steuerliche Behandlung belegt.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    audit = build_audit()
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(audit, indent=2, ensure_ascii=False), encoding="utf-8")
    OUT_MD.write_text(render_doc(audit), encoding="utf-8")
    print(f"Wrote {OUT_JSON.relative_to(ROOT)}")
    print(f"Wrote {OUT_MD.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
